"""Skill: 调用 LLM 生成英文例句和选词填空，写入 Excel。

所有 LLM prompt 存放在本 skill 的 prompts/ 目录中，自包含不依赖外部。
"""

import json
import random
import time
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

from lib.config import (
    LLM_API_KEY, LLM_BASE_URL, LLM_MODEL, MAX_RETRIES,
    GROUP_SIZE, VOCAB_SHEET, CLOZE_SHEET,
)

# 本 skill 自带的 prompt 目录
_PROMPT_DIR = Path(__file__).parent.parent / "prompts"


@dataclass
class ClozeQuestion:
    """一道选词填空题。"""
    sentence: str       # 含 _____ 的句子
    answer: str         # 正确答案
    chinese_hint: str   # 简短中文提示


@dataclass
class ClozeGroup:
    """一组选词填空（10 题）。"""
    group_id: int
    word_bank: list[str]         # 单词表（保持顺序）
    questions: list[ClozeQuestion]  # 题目（已乱序）


def run(words: list[dict], output_path: str) -> dict:
    """完整流程：生成例句 → 生成填空 → 写入 Excel。"""
    from openai import OpenAI
    client = OpenAI(api_key=LLM_API_KEY, base_url=LLM_BASE_URL)

    groups = [words[i:i + GROUP_SIZE] for i in range(0, len(words), GROUP_SIZE)]

    # 1. 生成例句（每批 8 个，减少 API 调用）
    print(f"  为 {len(words)} 个单词生成例句...")
    sentence_prompt = (_PROMPT_DIR / "sentence.md").read_text(encoding="utf-8")
    all_sentences = []
    for i in range(0, len(words), 8):
        batch = words[i:i + 8]
        prompt = sentence_prompt.replace("{word_table}", _build_table(batch))
        all_sentences.extend(_gen_sentences(client, prompt, batch))
        if i + 8 < len(words):
            time.sleep(0.5)
    print(f"  已生成 {len(all_sentences)} 条例句")

    # 2. 生成选词填空（每组 10 个单词）
    print(f"  为 {len(groups)} 组生成选词填空...")
    cloze_prompt = (_PROMPT_DIR / "cloze.md").read_text(encoding="utf-8")
    cloze_groups = []
    for gi, group in enumerate(groups, 1):
        prompt = cloze_prompt.replace("{word_table}", _build_table(group)).replace("{count}", str(len(group)))
        cg = _gen_cloze(client, prompt, group, gi)
        cloze_groups.append(cg)
        time.sleep(0.5)
    print(f"  已生成 {len(cloze_groups)} 组填空")

    # 3. 写入 Excel
    path = _write_excel(groups, all_sentences, cloze_groups, output_path)

    return {"filepath": str(path.resolve()), "words": len(words), "groups": len(groups)}


def _build_table(words: list[dict]) -> str:
    """构建 LLM prompt 中的单词表格（markdown 格式）。"""
    return "\n".join(
        "| {} | {} | {} | {} |".format(i, w["word"], w["pos"], w["translation"])
        for i, w in enumerate(words, 1)
    )


# ── LLM 调用 ──────────────────────────────────────────

def _gen_sentences(client, prompt: str, batch: list[dict]) -> list[str]:
    """调用 LLM 生成例句，失败时降级为模板句。"""
    for attempt in range(MAX_RETRIES):
        try:
            resp = client.chat.completions.create(
                model=LLM_MODEL,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=2048, temperature=0.7,
            )
            text = resp.choices[0].message.content.strip()
            if text.startswith("```"): text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
            data = json.loads(text)
            return [f"{item['sentence']}\n{item['translation']}" for item in data]
        except Exception:
            if attempt == MAX_RETRIES - 1:
                return [_fallback_sent(w) for w in batch]
            time.sleep(1)
    return []


def _gen_cloze(client, prompt: str, words: list[dict], group_id: int) -> ClozeGroup:
    """调用 LLM 生成选词填空，单词表有序、题目乱序。"""
    bank = [w["word"] for w in words]
    for attempt in range(MAX_RETRIES):
        try:
            resp = client.chat.completions.create(
                model=LLM_MODEL,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=2048, temperature=0.7,
            )
            text = resp.choices[0].message.content.strip()
            if text.startswith("```"): text = text.split("```")[1]
            if text.startswith("json"): text = text[4:]
            data = json.loads(text)
            qs = [ClozeQuestion(item["sentence"], item["word"], item.get("hint", "")) for item in data]
            random.shuffle(qs)
            return ClozeGroup(group_id, bank, qs)
        except Exception:
            if attempt == MAX_RETRIES - 1:
                return _fallback_cloze(words, group_id)
            time.sleep(1)
    return _fallback_cloze(words, group_id)


def _fallback_sent(w: dict) -> str:
    """例句生成失败时的备用模板。"""
    m = {"n.": "I know this {w}.", "v.": "I {w} it every day.",
         "adj.": "It is very {w}.", "adv.": "She did it {w}."}
    template = m.get(w["pos"].split("/")[0], "I learned the word {w}.")
    return template.format(w=w["word"]) + "\n（生成失败）"


def _fallback_cloze(words: list[dict], gid: int) -> ClozeGroup:
    """填空生成失败时的备用题目。"""
    qs = [ClozeQuestion("The word _____ is an example. ({})".format(w["pos"]), w["word"], w["translation"]) for w in words]
    random.shuffle(qs)
    return ClozeGroup(gid, [w["word"] for w in words], qs)


# ── Excel 输出 ────────────────────────────────────────

def _write_excel(word_groups, sentences, cloze_groups, output_path) -> Path:
    """写入纯文本 Excel：词汇表 Sheet + 选词填空 Sheet。"""
    output_path = Path(output_path)
    wb = Workbook()

    # Sheet 1: 词汇表
    ws = wb.active
    ws.title = VOCAB_SHEET
    for col, h in enumerate(["序号", "单词 词性 中文释义", "英文例句", "例句翻译", "组号"], 1):
        ws.cell(row=1, column=col, value=h)
    row, seq = 2, 1
    for gid, group in enumerate(word_groups, 1):
        for word in group:
            sf = sentences[seq - 1] if seq - 1 < len(sentences) else ""
            eng, chn = sf.split("\n", 1) if "\n" in sf else (sf, "")
            for col, val in enumerate([seq, f"{word['word']} {word['pos']} {word['translation']}", eng, chn, gid], 1):
                ws.cell(row=row, column=col, value=val)
            row += 1
            seq += 1
    for col, w in enumerate([6, 35, 45, 30, 6], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # Sheet 2: 选词填空
    ws = wb.create_sheet(CLOZE_SHEET)
    row = 1
    for cg in cloze_groups:
        ws.cell(row=row, column=1, value=f"第 {cg.group_id} 组")
        row += 1
        ws.cell(row=row, column=1, value=f"可选单词：{'，'.join(cg.word_bank)}")
        row += 1
        for col, h in enumerate(["题号", "题目（填空）", "答案", "中文提示"], 1):
            ws.cell(row=row, column=col, value=h)
        row += 1
        for qi, q in enumerate(cg.questions, 1):
            for col, val in enumerate([qi, q.sentence, q.answer, q.chinese_hint], 1):
                ws.cell(row=row, column=col, value=val)
            row += 1
        row += 1
    for col, w in enumerate([8, 55, 16, 18], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 保存（自动创建父目录）
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
